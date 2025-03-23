import Jetson.GPIO as GPIO
from datetime import datetime
import openpyxl
import time
import csv
import os

def setup_pins(input_pin, output_pin,led_input,led_output,relay_active):
    GPIO.setmode(GPIO.BOARD)
    GPIO.setup(input_pin, GPIO.OUT, initial=GPIO.HIGH)
    GPIO.setup(output_pin, GPIO.OUT, initial=GPIO.HIGH)
    GPIO.setup(led_input, GPIO.OUT, initial=GPIO.LOW)
    GPIO.setup(led_output, GPIO.OUT, initial=GPIO.LOW)
    GPIO.setup(relay_active, GPIO.OUT, initial=GPIO.HIGH)

def main():
    input_trip_pin = 11  # Input Relay Pin
    output_trip_pin = 12 # Output Relay Pin
    led_input = 15  # Input Trip Indication
    led_output = 16 # Output Trip Indication
    relay_active = 13 # Relay Active Indication

    setup_pins(input_trip_pin, output_trip_pin, led_input, led_output, relay_active)

def create_fault_log():
    """Creates an Excel and CSV file with predefined headers if they don't exist."""
    if not os.path.exists(fault_log_file):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Fault Log"
        sheet.append(["Timestamp", "Relay Status", "Input Status", "Output Status", "Circuit Breaker Status", "Fault Type"])
        workbook.save(fault_log_file)
        workbook.close()
        print(f"✅ Created fault log Excel file: {fault_log_file}")

    if not os.path.exists(fault_csv_file):
        with open(fault_csv_file, mode='w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["Timestamp", "Relay Status", "Input Status", "Output Status", "Circuit Breaker Status", "Fault Type"])
        print(f"✅ Created fault log CSV file: {fault_csv_file}")

def update_fault_log(relay_status, input_status, output_status, breaker_status, fault_type):
    """Updates Excel (real-time) and CSV (fault occurrences)."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 1️⃣ Update Excel (overwrite previous data)
    try:
        if os.path.exists(fault_log_file):
            workbook = openpyxl.load_workbook(fault_log_file)
            sheet = workbook.active
            # Clear all rows except header
            if sheet.max_row > 1:
                sheet.delete_rows(2, sheet.max_row - 1)
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Fault Log"
            sheet.append(["Timestamp", "Relay Status", "Input Status", "Output Status", "Circuit Breaker Status", "Fault Type"])
        
        # Append current status
        sheet.append([timestamp, relay_status, input_status, output_status, breaker_status, fault_type])
        workbook.save(fault_log_file)
        workbook.close()
    except Exception as e:
        print(f"Excel Update Error: {e}")

    # 2️⃣ Append to CSV only if there's a fault
    if fault_type != "None":
        try:
            with open(fault_csv_file, mode='a', newline='') as file:
                writer = csv.writer(file)
                writer.writerow([timestamp, relay_status, input_status, output_status, breaker_status, fault_type])
            print(f"✅ Logged fault: {fault_type} at {timestamp}")
        except Exception as e:
            print(f"CSV Append Error: {e}")


def read_excel_cells(file_path, sheet_name, cells):
    """Reads multiple cells from an Excel file and returns their values."""
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
       
        # Explicitly assign values
        input_phase_a_over_current_status = sheet[cells[0]].value  # B2
        input_phase_a_over_current_set_value = sheet[cells[1]].value  # B3
        input_phase_b_over_current_status = sheet[cells[2]].value  # B4
        input_phase_b_over_current_set_value = sheet[cells[3]].value
        input_phase_c_over_current_status = sheet[cells[4]].value
        input_phase_c_over_current_set_value = sheet[cells[5]].value
        input_phase_a_over_voltage_status = sheet[cells[6]].value
        input_phase_a_over_voltage_set_value = sheet[cells[7]].value
        input_phase_b_over_voltage_status = sheet[cells[8]].value
        input_phase_b_over_voltage_set_value = sheet[cells[9]].value
        input_phase_c_over_voltage_status = sheet[cells[10]].value
        input_phase_c_over_voltage_set_value = sheet[cells[11]].value
        input_phase_a_under_voltage_status = sheet[cells[12]].value
        input_phase_a_under_voltage_set_value = sheet[cells[13]].value
        input_phase_b_under_voltage_status = sheet[cells[14]].value
        input_phase_b_under_voltage_set_value = sheet[cells[15]].value
        input_phase_c_under_voltage_status = sheet[cells[16]].value
        input_phase_c_under_voltage_set_value = sheet[cells[17]].value
        input_over_frequency_status = sheet[cells[18]].value
        input_over_frequency_set_value = sheet[cells[19]].value
        input_under_frequency_status = sheet[cells[20]].value
        input_under_frequency_set_value = sheet[cells[21]].value
        input_dc_over_voltage_status = sheet[cells[22]].value
        input_dc_over_voltage_set_value = sheet[cells[23]].value
        input_dc_under_voltage_status = sheet[cells[24]].value
        input_dc_under_voltage_set_value = sheet[cells[25]].value
        input_dc_over_current_status = sheet[cells[26]].value
        input_dc_over_current_set_value = sheet[cells[27]].value
        input_over_temperature_status = sheet[cells[28]].value
        input_over_temperature_set_value = sheet[cells[29]].value
        output_phase_a_over_current_status = sheet[cells[30]].value
        output_phase_a_over_current_set_value = sheet[cells[31]].value  
        output_phase_b_over_current_status = sheet[cells[32]].value
        output_phase_b_over_current_set_value = sheet[cells[33]].value
        output_phase_c_over_current_status = sheet[cells[34]].value
        output_phase_c_over_current_set_value = sheet[cells[35]].value
        output_phase_a_over_voltage_status = sheet[cells[36]].value
        output_phase_a_over_voltage_set_value = sheet[cells[37]].value
        output_phase_b_over_voltage_status = sheet[cells[38]].value
        output_phase_b_over_voltage_set_value = sheet[cells[39]].value
        output_phase_c_over_voltage_status = sheet[cells[40]].value
        output_phase_c_over_voltage_set_value = sheet[cells[41]].value
        output_phase_a_under_voltage_status = sheet[cells[42]].value
        output_phase_a_under_voltage_set_value = sheet[cells[43]].value
        output_phase_b_under_voltage_status = sheet[cells[44]].value
        output_phase_b_under_voltage_set_value = sheet[cells[45]].value
        output_phase_c_under_voltage_status = sheet[cells[46]].value
        output_phase_c_under_voltage_set_value = sheet[cells[47]].value
        output_over_frequency_status = sheet[cells[48]].value
        output_over_frequency_set_value = sheet[cells[49]].value
        output_under_frequency_status = sheet[cells[50]].value
        output_under_frequency_set_value = sheet[cells[51]].value
        output_dc_over_voltage_status = sheet[cells[52]].value
        output_dc_over_voltage_set_value = sheet[cells[53]].value
        output_dc_under_voltage_status = sheet[cells[54]].value
        output_dc_under_voltage_set_value = sheet[cells[55]].value
        output_dc_over_current_status = sheet[cells[56]].value
        output_dc_over_current_set_value = sheet[cells[57]].value
        output_over_temperature_status = sheet[cells[58]].value
        output_over_temperature_set_value = sheet[cells[59]].value
        Instantaneous_Trip_Characteristics_status = sheet[cells[60]].value
        Inverse_Time_Characteristics_status = sheet[cells[61]].value
        Definite_Time_Characteristics_status = sheet[cells[62]].value
        Differential_Relay_Characteristics_status = sheet[cells[63]].value
        Trip_button = sheet[cells[64]].value
        Reset_button = sheet[cells[65]].value  


        workbook.close()
        return (input_phase_a_over_current_status, input_phase_a_over_current_set_value , 
                input_phase_b_over_current_status, input_phase_b_over_current_set_value,
                input_phase_c_over_current_status,input_phase_c_over_current_set_value,
                input_phase_a_over_voltage_status,input_phase_a_over_voltage_set_value ,
                input_phase_b_over_voltage_status,input_phase_b_over_voltage_set_value,
                input_phase_c_over_voltage_status ,input_phase_c_over_voltage_set_value,
                input_phase_a_under_voltage_status ,input_phase_a_under_voltage_set_value,
                input_phase_b_under_voltage_status,input_phase_b_under_voltage_set_value,
                input_phase_c_under_voltage_status,input_phase_c_under_voltage_set_value,
                input_over_frequency_status,input_over_frequency_set_value,input_under_frequency_status,
                input_under_frequency_set_value,input_dc_over_voltage_status,input_dc_over_voltage_set_value,
                input_dc_under_voltage_status,input_dc_under_voltage_set_value,input_dc_over_current_status,
                input_dc_over_current_set_value,input_over_temperature_status,input_over_temperature_set_value,
                output_phase_a_over_current_status,output_phase_a_over_current_set_value,output_phase_b_over_current_status,
                output_phase_b_over_current_set_value,output_phase_c_over_current_status,output_phase_c_over_current_set_value,
                output_phase_a_over_voltage_status,output_phase_a_over_voltage_set_value,output_phase_b_over_voltage_status,
                output_phase_b_over_voltage_set_value,output_phase_c_over_voltage_status,output_phase_c_over_voltage_set_value,
                output_phase_a_under_voltage_status,output_phase_a_under_voltage_set_value,output_phase_b_under_voltage_status,
                output_phase_b_under_voltage_set_value,output_phase_c_under_voltage_status,output_phase_c_under_voltage_set_value,
                output_over_frequency_status,output_over_frequency_set_value,output_under_frequency_status,output_under_frequency_set_value,
                output_dc_over_voltage_status,output_dc_over_voltage_set_value,output_dc_under_voltage_status,output_dc_under_voltage_set_value,
                output_dc_over_current_status,output_dc_over_current_set_value,output_over_temperature_status,output_over_temperature_set_value,
                Instantaneous_Trip_Characteristics_status,Inverse_Time_Characteristics_status,Definite_Time_Characteristics_status,
                Differential_Relay_Characteristics_status,Trip_button,Reset_button)


    except Exception as e:
        print(f"Excel Read Error: {e}")
        return (None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,
                None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,
                None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None)

def read_input_csv(file_path):
    """Reads values from the second row of the input CSV file."""
    try:
        with open(file_path, mode='r', newline='') as file:
            reader = csv.reader(file)
            next(reader)  # Skip the header row
            row = next(reader, None)  # Read the second row

            if row:
                # Assign input values explicitly
                input_Computer_TS = row[0]  # B2
                input_A_Phase_Voltage = row[1]  # C2
                input_A_Phase_Current = row[2]  # D2
                input_A_Phase_Active_Power = row[3]  # E2
                input_A_Phase_Reactive_Power = row[4]  # F2
                input_A_Phase_Apparent_Power = row[5]  # G2
                input_A_Power_Factor = row[6]  # H2
                input_B_Phase_Voltage = row[7]  # I2
                input_B_Phase_Current = row[8]  # J2
                input_B_Phase_Active_Power = row[9]  # K2
                input_B_Phase_Reactive_Power = row[10]  # L2
                input_B_Phase_Apparent_Power = row[11]  # M2
                input_B_Power_Factor = row[12]  # N2
                input_C_Phase_Voltage = row[13]  # O2
                input_C_Phase_Current = row[14]  # P2
                input_C_Phase_Active_Power = row[15]  # Q2
                input_C_Phase_Reactive_Power = row[16]  # R2
                input_C_Phase_Apparent_Power = row[17]  # S2
                input_C_Power_Factor = row[18]  # T2
                input_Frequency = row[19]  # U2
                input_DC_Voltage = row[20]  # V2
                input_DC_Current = row[21]  # W2
                input_Temperature = row[22]  # X2

                return (input_Computer_TS, input_A_Phase_Voltage, input_A_Phase_Current, input_A_Phase_Active_Power, 
                        input_A_Phase_Reactive_Power, input_A_Phase_Apparent_Power, input_A_Power_Factor, 
                        input_B_Phase_Voltage, input_B_Phase_Current, input_B_Phase_Active_Power, 
                        input_B_Phase_Reactive_Power, input_B_Phase_Apparent_Power, input_B_Power_Factor, 
                        input_C_Phase_Voltage, input_C_Phase_Current, input_C_Phase_Active_Power, 
                        input_C_Phase_Reactive_Power, input_C_Phase_Apparent_Power, input_C_Power_Factor, 
                        input_Frequency, input_DC_Voltage, input_DC_Current, input_Temperature)

            else:
                return (None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                         None, None, None, None, None, None, None, None)

    except Exception as e:
        print(f"Input CSV Read Error: {e}")
        return None, None, None, None

def read_output_csv(file_path):
    """Reads values from the second row of the output CSV file."""
    try:
        with open(file_path, mode='r', newline='') as file:
            reader = csv.reader(file)
            next(reader)  # Skip the header row
            row = next(reader, None)  # Read the second row

            if row:
                # Assign output values explicitly
                output_Computer_TS = row[0]  # B2
                output_A_Phase_Voltage = row[1]  # C2
                output_A_Phase_Current = row[2]  # D2
                output_A_Phase_Active_Power = row[3]  # E2
                output_A_Phase_Reactive_Power = row[4]  # F2
                output_A_Phase_Apparent_Power = row[5]  # G2
                output_A_Power_Factor = row[6]  # H2
                output_B_Phase_Voltage = row[7]  # I2
                output_B_Phase_Current = row[8]  # J2
                output_B_Phase_Active_Power = row[9]  # K2
                output_B_Phase_Reactive_Power = row[10]  # L2
                output_B_Phase_Apparent_Power = row[11]  # M2
                output_B_Power_Factor = row[12]  # N2
                output_C_Phase_Voltage = row[13]  # O2
                output_C_Phase_Current = row[14]  # P2
                output_C_Phase_Active_Power = row[15]  # Q2
                output_C_Phase_Reactive_Power = row[16]  # R2
                output_C_Phase_Apparent_Power = row[17]  # S2
                output_C_Power_Factor = row[18]  # T2
                output_Frequency = row[19]  # U2
                output_DC_Voltage = row[20]  # V2
                output_DC_Current = row[21]  # W2
                output_Temperature = row[22]  # X2

                return (output_Computer_TS, output_A_Phase_Voltage, output_A_Phase_Current, output_A_Phase_Active_Power, 
                        output_A_Phase_Reactive_Power, output_A_Phase_Apparent_Power, output_A_Power_Factor, 
                        output_B_Phase_Voltage, output_B_Phase_Current, output_B_Phase_Active_Power, 
                        output_B_Phase_Reactive_Power, output_B_Phase_Apparent_Power, output_B_Power_Factor, 
                        output_C_Phase_Voltage, output_C_Phase_Current, output_C_Phase_Active_Power, 
                        output_C_Phase_Reactive_Power, output_C_Phase_Apparent_Power, output_C_Power_Factor, 
                        output_Frequency, output_DC_Voltage, output_DC_Current, output_Temperature)
            else:
                return (None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                         None, None, None, None, None, None, None, None)

    except Exception as e:
        print(f"Output CSV Read Error: {e}")
        return None, None, None, None

def monitor_files(excel_path, sheet_name, excel_cells, input_csv, output_csv, input_trip_pin=11, output_trip_pin=13, interval=1):
    """Continuously reads Excel and CSV files, updates variables, and performs checks."""
    input_trip_pin = 11  # Input Relay Pin
    output_trip_pin = 12 # Output Relay Pin
    led_input = 15  # Input Trip Indication
    led_output = 16 # Output Trip Indication
    relay_active = 13 # Relay Active Indication

    setup_pins(input_trip_pin, output_trip_pin, led_input, led_output, relay_active)

    try:
        while True:
            # Read Excel file
            (input_phase_a_over_current_status, input_phase_a_over_current_set_value , input_phase_b_over_current_status, 
             input_phase_b_over_current_set_value,input_phase_c_over_current_status,input_phase_c_over_current_set_value,
             input_phase_a_over_voltage_status,input_phase_a_over_voltage_set_value ,input_phase_b_over_voltage_status,
             input_phase_b_over_voltage_set_value,input_phase_c_over_voltage_status ,input_phase_c_over_voltage_set_value,
             input_phase_a_under_voltage_status ,input_phase_a_under_voltage_set_value, input_phase_b_under_voltage_status,
             input_phase_b_under_voltage_set_value,input_phase_c_under_voltage_status,input_phase_c_under_voltage_set_value,
             input_over_frequency_status,input_over_frequency_set_value,input_under_frequency_status,input_under_frequency_set_value,
             input_dc_over_voltage_status,input_dc_over_voltage_set_value,input_dc_under_voltage_status,input_dc_under_voltage_set_value,
             input_dc_over_current_status,input_dc_over_current_set_value,input_over_temperature_status,input_over_temperature_set_value,
             output_phase_a_over_current_status,output_phase_a_over_current_set_value,output_phase_b_over_current_status,output_phase_b_over_current_set_value,
             output_phase_c_over_current_status,output_phase_c_over_current_set_value,output_phase_a_over_voltage_status,output_phase_a_over_voltage_set_value,
             output_phase_b_over_voltage_status,output_phase_b_over_voltage_set_value,output_phase_c_over_voltage_status,output_phase_c_over_voltage_set_value,
             output_phase_a_under_voltage_status,output_phase_a_under_voltage_set_value,output_phase_b_under_voltage_status,output_phase_b_under_voltage_set_value,
             output_phase_c_under_voltage_status,output_phase_c_under_voltage_set_value,output_over_frequency_status,output_over_frequency_set_value,output_under_frequency_status,
             output_under_frequency_set_value,output_dc_over_voltage_status,output_dc_over_voltage_set_value,output_dc_under_voltage_status,output_dc_under_voltage_set_value,
             output_dc_over_current_status,output_dc_over_current_set_value,output_over_temperature_status,output_over_temperature_set_value, Instantaneous_Trip_Characteristics_status,
             Inverse_Time_Characteristics_status,Definite_Time_Characteristics_status,Differential_Relay_Characteristics_status,Trip_button,Reset_button) = read_excel_cells(excel_path, sheet_name, excel_cells)
            
            # Read Input CSV file
            (input_Computer_TS, input_A_Phase_Voltage, input_A_Phase_Current, input_A_Phase_Active_Power, 
            input_A_Phase_Reactive_Power, input_A_Phase_Apparent_Power, input_A_Power_Factor, 
            input_B_Phase_Voltage, input_B_Phase_Current, input_B_Phase_Active_Power, 
            input_B_Phase_Reactive_Power, input_B_Phase_Apparent_Power, input_B_Power_Factor, 
            input_C_Phase_Voltage, input_C_Phase_Current, input_C_Phase_Active_Power, 
            input_C_Phase_Reactive_Power, input_C_Phase_Apparent_Power, input_C_Power_Factor, 
            input_Frequency, input_DC_Voltage, input_DC_Current, input_Temperature) = read_input_csv(input_csv)
            
            # Read Output CSV file
            (output_Computer_TS, output_A_Phase_Voltage, output_A_Phase_Current, output_A_Phase_Active_Power, 
            output_A_Phase_Reactive_Power, output_A_Phase_Apparent_Power, output_A_Power_Factor, 
            output_B_Phase_Voltage, output_B_Phase_Current, output_B_Phase_Active_Power, 
            output_B_Phase_Reactive_Power, output_B_Phase_Apparent_Power, output_B_Power_Factor, 
            output_C_Phase_Voltage, output_C_Phase_Current, output_C_Phase_Active_Power, 
            output_C_Phase_Reactive_Power, output_C_Phase_Apparent_Power, output_C_Power_Factor, 
            output_Frequency, output_DC_Voltage, output_DC_Current, output_Temperature) = read_output_csv(output_csv)

            # Print the extracted values for debuging
            #print(f"Excel Data: input_phase_a_over_current_status={input_phase_b_over_current_status}")
            #print(f"Input CSV Data: input1={input_B_Phase_Current}, input_phase_a_over_current_set_value={input_Temperature}, input3={input_DC_Current}, input4={input_B_Phase_Voltage}")
            #print(f"Output CSV Data: output1={output_A_Phase_Voltage}, output2={output_B_Phase_Apparent_Power}, output3={output_Temperature}, output4={output_DC_Voltage}")
            
                                    # Default healthy values (MISSING - ADD THIS)
            relay_status = "Healthy and Operational"
            input_status = "Healthy"
            output_status = "Healthy"
            breaker_status = "Live"
            fault_type = "None"
            

  
            if Trip_button == 1:
                print("trip Button Activated")
                GPIO.output(input_trip_pin, GPIO.LOW)
                GPIO.output(output_trip_pin, GPIO.LOW)
                GPIO.output(led_input, GPIO.HIGH)
                GPIO.output(led_output, GPIO.HIGH)
                relay_status = "Unhealthy"    # ✅ Add these lines
                breaker_status = "Trip"
                fault_type = "Manual Trip"

               # ========================= INTPUT RELAY LOGIC FOR INSTANTANEOUS TRIPING =========================
             
            if Instantaneous_Trip_Characteristics_status == 1:
               
               # Input Phase A Overcurrent
               if input_phase_a_over_current_status == 1:
                  
                  if input_A_Phase_Current is not None and input_phase_a_over_current_set_value is not None:
                    try:
                        input_A_Phase_Current = float(input_A_Phase_Current)
                        input_phase_a_over_current_set_value = float(input_phase_a_over_current_set_value)

                        if input_A_Phase_Current >= input_phase_a_over_current_set_value:
                           print("trip on - Input Phase A Overcurrent")
                           GPIO.output(input_trip_pin, GPIO.LOW)
                           GPIO.output(led_input, GPIO.HIGH)
                           input_status = "Unhealthy"
                           breaker_status = "Trip"
                           fault_type = "Input Phase A Overcurrent"
                        
                    except ValueError as e:
                       print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                # Input Phase B Overcurrent
               if input_phase_b_over_current_status == 1:
               
                   if input_B_Phase_Current is not None and input_phase_b_over_current_set_value is not None:
                   
                    try:
                       input_B_Phase_Current = float(input_B_Phase_Current)
                       input_phase_b_over_current_set_value = float(input_phase_b_over_current_set_value)

                       if input_B_Phase_Current >= input_phase_b_over_current_set_value:
                          print("trip on - Input Phase B Overcurrent")
                          GPIO.output(input_trip_pin, GPIO.LOW)
                          GPIO.output(led_input, GPIO.HIGH)
                          input_status = "Unhealthy"
                          breaker_status = "Trip"
                          fault_type = "Input Phase B Overcurrent"
                        
                    except ValueError as e:
                        print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                # Input Phase C Overcurrent
               if input_phase_c_over_current_status == 1:
                  if input_C_Phase_Current is not None and input_phase_c_over_current_set_value is not None:
                    try:
                       input_C_Phase_Current = float(input_C_Phase_Current)
                       input_phase_c_over_current_set_value = float(input_phase_c_over_current_set_value)
                       if input_C_Phase_Current >= input_phase_c_over_current_set_value:
                          print("trip on - Input Phase C Overcurrent")
                          GPIO.output(input_trip_pin, GPIO.LOW)
                          GPIO.output(led_input, GPIO.HIGH)
                          input_status = "Unhealthy"
                          breaker_status = "Trip"
                          fault_type = "Input Phase C Overcurrent"
                        
                    except ValueError as e:
                      print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                # Input DC Over Current
               if input_dc_over_current_status == 1:
                  if input_DC_Current is not None and input_dc_over_current_set_value is not None:
                     try:
                       input_DC_Current = float(input_DC_Current)
                       input_dc_over_current_set_value = float(input_dc_over_current_set_value)
                       if input_DC_Current >= input_dc_over_current_set_value:
                         print("trip on - Input DC Over Current")
                         GPIO.output(input_trip_pin, GPIO.LOW)
                         GPIO.output(led_input, GPIO.HIGH)
                         input_status = "Unhealthy"
                         breaker_status = "Trip"
                         fault_type = "Input DC Overcurrent"
                        
                     except ValueError as e:
                        print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                    # Input Over Temperature
               if input_over_temperature_status == 1:
                  if input_Temperature is not None and input_over_temperature_set_value is not None:
                     try:
                        input_Temperature = float(input_Temperature)
                        input_over_temperature_set_value = float(input_over_temperature_set_value)
                        if input_Temperature >= input_over_temperature_set_value:
                           print("trip on - Input Over Temperature")
                           GPIO.output(input_trip_pin, GPIO.LOW)
                           GPIO.output(led_input, GPIO.HIGH)
                           input_status = "Unhealthy"
                           breaker_status = "Trip"
                           fault_type = "Input Over Temperature"
                        
                     except ValueError as e:
                           print(f"Conversion Error: {e} - Check if input values are valid numbers.")

               # Phase A Overvoltage
               if input_phase_a_over_voltage_status == 1:
                  if input_A_Phase_Voltage is not None and input_phase_a_over_voltage_set_value is not None:
                     try:
                        input_A_Phase_Voltage = float(input_A_Phase_Voltage)
                        input_phase_a_over_voltage_set_value = float(input_phase_a_over_voltage_set_value)
                        if input_A_Phase_Voltage >= input_phase_a_over_voltage_set_value:
                          print("trip on - Input Phase A Overvoltage")
                          GPIO.output(input_trip_pin, GPIO.LOW)
                          GPIO.output(led_input, GPIO.HIGH)
                          input_status = "Unhealthy"
                          breaker_status = "Trip"
                          fault_type = "Input Phase A Overvoltage"
                        
                     except ValueError as e:
                        print(f"Conversion Error: {e} - Check if input values are valid numbers.")

               # Phase B Overvoltage
               if input_phase_b_over_voltage_status == 1:
                  if input_B_Phase_Voltage is not None and input_phase_b_over_voltage_set_value is not None:
                     try:
                       input_B_Phase_Voltage = float(input_B_Phase_Voltage)
                       input_phase_b_over_voltage_set_value = float(input_phase_b_over_voltage_set_value)
                       if input_B_Phase_Voltage >= input_phase_b_over_voltage_set_value:
                          print("trip on - Input Phase B Overvoltage")
                          GPIO.output(input_trip_pin, GPIO.LOW)
                          GPIO.output(led_input, GPIO.HIGH)
                          input_status = "Unhealthy"
                          breaker_status = "Trip"
                          fault_type = "Input Phase B Overvoltage"
                       
                     except ValueError as e:
                          print(f"Conversion Error: {e} - Check if input values are valid numbers.")

               # Phase C Overvoltage
               if input_phase_c_over_voltage_status == 1:
                  if input_C_Phase_Voltage is not None and input_phase_c_over_voltage_set_value is not None:
                     try:
                        input_C_Phase_Voltage = float(input_C_Phase_Voltage)
                        input_phase_c_over_voltage_set_value = float(input_phase_c_over_voltage_set_value)
                        if input_C_Phase_Voltage >= input_phase_c_over_voltage_set_value:
                          print("trip on - Input Phase C Overvoltage")
                          GPIO.output(input_trip_pin, GPIO.LOW)
                          GPIO.output(led_input, GPIO.HIGH)
                          input_status = "Unhealthy"
                          breaker_status = "Trip"
                          fault_type = "Input Phase C Overvoltage"
                        
                     except ValueError as e:
                          print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                # Phase A Undervoltage
               if input_phase_a_under_voltage_status == 1:
                  if input_A_Phase_Voltage is not None and input_phase_a_under_voltage_set_value is not None:
                     try:
                        input_A_Phase_Voltage = float(input_A_Phase_Voltage)
                        input_phase_a_under_voltage_set_value = float(input_phase_a_under_voltage_set_value)
                        if input_A_Phase_Voltage <= input_phase_a_under_voltage_set_value:
                           print("trip on - Input Phase A Undervoltage")
                           GPIO.output(input_trip_pin, GPIO.LOW)
                           GPIO.output(led_input, GPIO.HIGH)
                           input_status = "Unhealthy"
                           breaker_status = "Trip"
                           fault_type = "Input Phase A Undervoltage"
                        
                     except ValueError as e:
                        print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                # Phase B Undervoltage
               if input_phase_b_under_voltage_status == 1:
                  if input_B_Phase_Voltage is not None and input_phase_b_under_voltage_set_value is not None:
                     try:
                        input_B_Phase_Voltage = float(input_B_Phase_Voltage)
                        input_phase_b_under_voltage_set_value = float(input_phase_b_under_voltage_set_value)
                        if input_B_Phase_Voltage <= input_phase_b_under_voltage_set_value:
                           print("trip on - Input Phase B Undervoltage")
                           GPIO.output(input_trip_pin, GPIO.LOW)
                           GPIO.output(led_input, GPIO.HIGH)
                           input_status = "Unhealthy"
                           breaker_status = "Trip"
                           fault_type = "Input Phase B Undervoltage"
                        
                     except ValueError as e:
                          print(f"Conversion Error: {e} - Check if input values are valid numbers.")

               # Phase C Undervoltage
               if input_phase_c_under_voltage_status == 1:
                  if input_C_Phase_Voltage is not None and input_phase_c_under_voltage_set_value is not None:
                     try:
                        input_C_Phase_Voltage = float(input_C_Phase_Voltage)
                        input_phase_c_under_voltage_set_value = float(input_phase_c_under_voltage_set_value)
                        if input_C_Phase_Voltage <= input_phase_c_under_voltage_set_value:
                          print("trip on - Input Phase C Undervoltage")
                          GPIO.output(input_trip_pin, GPIO.LOW)
                          GPIO.output(led_input, GPIO.HIGH)
                          input_status = "Unhealthy"
                          breaker_status = "Trip"
                          fault_type = "Input Phase C Undervoltage"
                        
                     except ValueError as e:
                          print(f"Conversion Error: {e} - Check if input values are valid numbers.")

               # Input DC Over Voltage
               if input_dc_over_voltage_status == 1:
                  if input_DC_Voltage is not None and input_dc_over_voltage_set_value is not None:
                     try:
                       input_DC_Voltage = float(input_DC_Voltage)
                       input_dc_over_voltage_set_value = float(input_dc_over_voltage_set_value)
                       if input_DC_Voltage >= input_dc_over_voltage_set_value:
                         print("trip on - Input DC Over Voltage")
                         GPIO.output(input_trip_pin, GPIO.LOW)
                         GPIO.output(led_input, GPIO.HIGH)
                         input_status = "Unhealthy"
                         breaker_status = "Trip"
                         fault_type = "Input DC Over Voltage"
                       
                     except ValueError as e:
                        print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                # Input DC Under Voltage
               if input_dc_under_voltage_status == 1:
                  if input_DC_Voltage is not None and input_dc_under_voltage_set_value is not None:
                     try:
                       input_DC_Voltage = float(input_DC_Voltage)
                       input_dc_under_voltage_set_value = float(input_dc_under_voltage_set_value)
                       if input_DC_Voltage <= input_dc_under_voltage_set_value:
                         print("trip on - Input DC Under Voltage")
                         GPIO.output(input_trip_pin, GPIO.LOW)
                         GPIO.output(led_input, GPIO.HIGH)
                         input_status = "Unhealthy"
                         breaker_status = "Trip"
                         fault_type = "Input DC Under Voltage"
                       
                     except ValueError as e:
                        print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                # Over Frequency
               if input_over_frequency_status == 1:
                  if input_Frequency is not None and input_over_frequency_set_value is not None:
                     try:
                        input_Frequency = float(input_Frequency)
                        input_over_frequency_set_value = float(input_over_frequency_set_value)
                        if input_Frequency >= input_over_frequency_set_value:
                           print("trip on - Input Over Frequency")
                           GPIO.output(input_trip_pin, GPIO.LOW)
                           GPIO.output(led_input, GPIO.HIGH)
                           input_status = "Unhealthy"
                           breaker_status = "Trip"
                           fault_type = "Input Over Frequency"
                        
                     except ValueError as e:
                           print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                # Under Frequency
               if input_under_frequency_status == 1:
                  if input_Frequency is not None and input_under_frequency_set_value is not None:
                     try:
                        input_Frequency = float(input_Frequency)
                        input_under_frequency_set_value = float(input_under_frequency_set_value)
                        if input_Frequency <= input_under_frequency_set_value:
                           print("trip on - Input Under Frequency")
                           GPIO.output(input_trip_pin, GPIO.LOW)
                           GPIO.output(led_input, GPIO.HIGH)
                           input_status = "Unhealthy"
                           breaker_status = "Trip"
                           fault_type = "Input Under Frequency"
                        
                     except ValueError as e:
                         print(f"Conversion Error: {e} - Check if input values are valid numbers.")


             # ========================= OUTPUT RELAY LOGIC FOR INSTANTANEOUS TRIPING =========================

                #   Output Phase A Overcurrent
               if output_phase_a_over_current_status == 1:
                  if output_A_Phase_Current is not None and output_phase_a_over_current_set_value is not None:
                     try:
                         output_A_Phase_Current = float(output_A_Phase_Current)
                         output_phase_a_over_current_set_value = float(output_phase_a_over_current_set_value)
                         if output_A_Phase_Current >= output_phase_a_over_current_set_value:
                            print("trip on - Output Phase A Overcurrent")
                            GPIO.output(output_trip_pin, GPIO.LOW)
                            GPIO.output(led_output, GPIO.HIGH)
                            output_status = "Unhealthy"
                            breaker_status = "Trip"
                            fault_type = "Output Phase A Overcurrent"
                         
                     except ValueError as e:
                            print(f"Conversion Error: {e} - Check if output values are valid numbers.")

                # Phase B Overcurrent
               if output_phase_b_over_current_status == 1:
                  if output_B_Phase_Current is not None and output_phase_b_over_current_set_value is not None:
                     try:
                         output_B_Phase_Current = float(output_B_Phase_Current)
                         output_phase_b_over_current_set_value = float(output_phase_b_over_current_set_value)
                         if output_B_Phase_Current >= output_phase_b_over_current_set_value:
                            print("trip on - Output Phase B Overcurrent")
                            GPIO.output(output_trip_pin, GPIO.LOW)
                            GPIO.output(led_output, GPIO.HIGH)
                            output_status = "Unhealthy"
                            breaker_status = "Trip"
                            fault_type = "Output Phase B Overcurrent"
                         
                     except ValueError as e:
                          print(f"Conversion Error: {e} - Check if output values are valid numbers.")

                # Phase C Overcurrent
               if output_phase_c_over_current_status == 1:
                  if output_C_Phase_Current is not None and output_phase_c_over_current_set_value is not None:
                      try:
                         output_C_Phase_Current = float(output_C_Phase_Current)
                         output_phase_c_over_current_set_value = float(output_phase_c_over_current_set_value)
                         if output_C_Phase_Current >= output_phase_c_over_current_set_value:
                            print("trip on - Output Phase C Overcurrent")
                            GPIO.output(output_trip_pin, GPIO.LOW)
                            GPIO.output(led_output, GPIO.HIGH)
                            output_status = "Unhealthy"
                            breaker_status = "Trip"
                            fault_type = "Output Phase C Overcurrent"
                         
                      except ValueError as e:
                           print(f"Conversion Error: {e} - Check if output values are valid numbers.")
 
                # DC Over Current
               if output_dc_over_current_status == 1:
                  if output_DC_Current is not None and output_dc_over_current_set_value is not None:
                     try:
                         output_DC_Current = float(output_DC_Current)
                         output_dc_over_current_set_value = float(output_dc_over_current_set_value)
                         if output_DC_Current >= output_dc_over_current_set_value:
                            print("trip on - Output DC Over Current")
                            GPIO.output(output_trip_pin, GPIO.LOW)
                            GPIO.output(led_output, GPIO.HIGH)
                            output_status = "Unhealthy"
                            breaker_status = "Trip"
                            fault_type = "Output DC Over Current"
                         
                     except ValueError as e:
                            print(f"Conversion Error: {e} - Check if output values are valid numbers.")

                # Over Temperature
               if output_over_temperature_status == 1:
                  if output_Temperature is not None and output_over_temperature_set_value is not None:
                     try:
                        output_Temperature = float(output_Temperature)
                        output_over_temperature_set_value = float(output_over_temperature_set_value)
                        if output_Temperature >= output_over_temperature_set_value:
                           print("trip on - Output Over Temperature")
                           GPIO.output(output_trip_pin, GPIO.LOW)
                           GPIO.output(led_output, GPIO.HIGH)
                           output_status = "Unhealthy"
                           breaker_status = "Trip"
                           fault_type = "Output Over Temperature"
                        
                     except ValueError as e:
                            print(f"Conversion Error: {e} - Check if output values are valid numbers.")


                # Phase A Overvoltage
               if output_phase_a_over_voltage_status == 1:
                  if output_A_Phase_Voltage is not None and output_phase_a_over_voltage_set_value is not None:
                     try:
                        output_A_Phase_Voltage = float(output_A_Phase_Voltage)
                        output_phase_a_over_voltage_set_value = float(output_phase_a_over_voltage_set_value)
                        if output_A_Phase_Voltage >= output_phase_a_over_voltage_set_value:
                           print("trip on - Output Phase A Overvoltage")
                           GPIO.output(output_trip_pin, GPIO.LOW)
                           GPIO.output(led_output, GPIO.HIGH)
                           output_status = "Unhealthy"
                           breaker_status = "Trip"
                           fault_type = "Output Phase A Overvoltage"
                        
                     except ValueError as e:
                          print(f"Conversion Error: {e} - Check if output values are valid numbers.")

                # Phase B Overvoltage
               if output_phase_b_over_voltage_status == 1:
                   if output_B_Phase_Voltage is not None and output_phase_b_over_voltage_set_value is not None:
                      try:
                          output_B_Phase_Voltage = float(output_B_Phase_Voltage)
                          output_phase_b_over_voltage_set_value = float(output_phase_b_over_voltage_set_value)
                          if output_B_Phase_Voltage >= output_phase_b_over_voltage_set_value:
                            print("trip on - Output Phase B Overvoltage")
                            GPIO.output(output_trip_pin, GPIO.LOW)
                            GPIO.output(led_output, GPIO.HIGH)
                            output_status = "Unhealthy"
                            breaker_status = "Trip"
                            fault_type = "Output Phase B Overvoltage"
                          
                      except ValueError as e:
                           print(f"Conversion Error: {e} - Check if output values are valid numbers.")

                # Phase C Overvoltage
               if output_phase_c_over_voltage_status == 1:
                   if output_C_Phase_Voltage is not None and output_phase_c_over_voltage_set_value is not None:
                      try:
                          output_C_Phase_Voltage = float(output_C_Phase_Voltage)
                          output_phase_c_over_voltage_set_value = float(output_phase_c_over_voltage_set_value)
                          if output_C_Phase_Voltage >= output_phase_c_over_voltage_set_value:
                             print("trip on - Output Phase C Overvoltage")
                             GPIO.output(output_trip_pin, GPIO.LOW)
                             GPIO.output(led_output, GPIO.HIGH)
                             output_status = "Unhealthy"
                             breaker_status = "Trip"
                             fault_type = "Output Phase C Overvoltage"
                         
                      except ValueError as e:
                              print(f"Conversion Error: {e} - Check if output values are valid numbers.")

                # Phase A Undervoltage
               if output_phase_a_under_voltage_status == 1:
                  if output_A_Phase_Voltage is not None and output_phase_a_under_voltage_set_value is not None:
                     try:
                         output_A_Phase_Voltage = float(output_A_Phase_Voltage)
                         output_phase_a_under_voltage_set_value = float(output_phase_a_under_voltage_set_value)
                         if output_A_Phase_Voltage <= output_phase_a_under_voltage_set_value:
                               print("trip on - Output Phase A Undervoltage")
                               GPIO.output(output_trip_pin, GPIO.LOW)
                               GPIO.output(led_output, GPIO.HIGH)
                               output_status = "Unhealthy"
                               breaker_status = "Trip"
                               fault_type = "Output Phase A Undervoltage"
                        
                     except ValueError as e:
                              print(f"Conversion Error: {e} - Check if output values are valid numbers.")

                # Phase B Undervoltage
               if output_phase_b_under_voltage_status == 1:
                   if output_B_Phase_Voltage is not None and output_phase_b_under_voltage_set_value is not None:
                      try:
                         output_B_Phase_Voltage = float(output_B_Phase_Voltage)
                         output_phase_b_under_voltage_set_value = float(output_phase_b_under_voltage_set_value)
                         if output_B_Phase_Voltage <= output_phase_b_under_voltage_set_value:
                            print("trip on - Output Phase B Undervoltage")
                            GPIO.output(output_trip_pin, GPIO.LOW)
                            GPIO.output(led_output, GPIO.HIGH)
                            output_status = "Unhealthy"
                            breaker_status = "Trip"
                            fault_type = "Output Phase B Undervoltage"
                        
                      except ValueError as e:
                            print(f"Conversion Error: {e} - Check if output values are valid numbers.")

               # Phase C Undervoltage
               if output_phase_c_under_voltage_status == 1:
                  if output_C_Phase_Voltage is not None and output_phase_c_under_voltage_set_value is not None:
                     try:
                         output_C_Phase_Voltage = float(output_C_Phase_Voltage)
                         output_phase_c_under_voltage_set_value = float(output_phase_c_under_voltage_set_value)
                         if output_C_Phase_Voltage <= output_phase_c_under_voltage_set_value:
                            print("trip on - Output Phase C Undervoltage")
                            GPIO.output(output_trip_pin, GPIO.LOW)
                            GPIO.output(led_output, GPIO.HIGH)
                            output_status = "Unhealthy"
                            breaker_status = "Trip"
                            fault_type = "Output Phase C Undervoltage"
                        
                     except ValueError as e:
                        print(f"Conversion Error: {e} - Check if output values are valid numbers.")

               # Output DC Over Voltage
               if output_dc_over_voltage_status == 1:
                  if output_DC_Voltage is not None and output_dc_over_voltage_set_value is not None:
                     try:
                       output_DC_Voltage = float(output_DC_Voltage)
                       output_dc_over_voltage_set_value = float(output_dc_over_voltage_set_value)
                       if output_DC_Voltage >= output_dc_over_voltage_set_value:
                         print("trip on - Output DC Over Voltage")
                         GPIO.output(output_trip_pin, GPIO.LOW)
                         GPIO.output(led_output, GPIO.HIGH)
                         output_status = "Unhealthy"
                         breaker_status = "Trip"
                         fault_type = "Output DC Over Voltage"
                       
                     except ValueError as e:
                        print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                # Output DC Under Voltage
               if output_dc_under_voltage_status == 1:
                  if output_DC_Voltage is not None and output_dc_under_voltage_set_value is not None:
                     try:
                       output_DC_Voltage = float(output_DC_Voltage)
                       output_dc_under_voltage_set_value = float(output_dc_under_voltage_set_value)
                       if output_DC_Voltage <= output_dc_under_voltage_set_value:
                         print("trip on - Output DC Under Voltage")
                         GPIO.output(output_trip_pin, GPIO.LOW)
                         GPIO.output(led_output, GPIO.HIGH)
                         output_status = "Unhealthy"
                         breaker_status = "Trip"
                         fault_type = "Output DC Under Voltage"
                       
                     except ValueError as e:
                        print(f"Conversion Error: {e} - Check if input values are valid numbers.")


                # Over Frequency
               if output_over_frequency_status == 1:
                  if output_Frequency is not None and output_over_frequency_set_value is not None:
                     try:
                        output_Frequency = float(output_Frequency)
                        output_over_frequency_set_value = float(output_over_frequency_set_value)
                        if output_Frequency >= output_over_frequency_set_value:
                           print("trip on - Output Over Frequency")
                           GPIO.output(output_trip_pin, GPIO.LOW)
                           GPIO.output(led_output, GPIO.HIGH)
                           output_status = "Unhealthy"
                           breaker_status = "Trip"
                           fault_type = "Output Over Frequency"
                        
                     except ValueError as e:
                          print(f"Conversion Error: {e} - Check if output values are valid numbers.")

                # Under Frequency
               if output_under_frequency_status == 1:
                  if output_Frequency is not None and output_under_frequency_set_value is not None:
                     try:
                         output_Frequency = float(output_Frequency)
                         output_under_frequency_set_value = float(output_under_frequency_set_value)
                         if output_Frequency <= output_under_frequency_set_value:
                            print("trip on - Output Under Frequency")
                            GPIO.output(output_trip_pin, GPIO.LOW)
                            GPIO.output(led_output, GPIO.HIGH)
                            output_status = "Unhealthy"
                            breaker_status = "Trip"
                            fault_type = "Output Under Frequency"
                         
                     except ValueError as e:
                          print(f"Conversion Error: {e} - Check if output values are valid numbers.")

                 # ========================= INVERSE TIME OVERCURRENT LOGIC =========================

            if Inverse_Time_Characteristics_status == 1:
                 # Phase A Overcurrent
              if input_phase_a_over_current_status == 1:
                
                 if input_A_Phase_Current is not None and input_phase_a_over_current_set_value is not None:
                     try:
                         input_A_Phase_Current = float(input_A_Phase_Current)
                         input_phase_a_over_current_set_value = float(input_phase_a_over_current_set_value)

                         if input_A_Phase_Current >= 20 * input_phase_a_over_current_set_value:
                             print("Input Phase A Overcurrent (20x) detected! Waiting 0.5 sec before trip.")
                             time.sleep(0.5)
                             GPIO.output(input_trip_pin, GPIO.LOW)
                             GPIO.output(led_input, GPIO.HIGH)
                             print("trip on - Input Phase A Overcurrent")
                             input_status = "Unhealthy"
                             breaker_status = "Trip"
                             fault_type = "Input Phase A Overcurrent"

                         elif input_A_Phase_Current >= 10 * input_phase_a_over_current_set_value:
                              print("Input Phase A Overcurrent (10x) detected! Waiting 1 sec before trip.")
                              time.sleep(1)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase A Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase A Overcurrent"

                         elif input_A_Phase_Current >= 5 * input_phase_a_over_current_set_value:
                              print("Input Phase A Overcurrent (5x) detected! Waiting 2.5 sec before trip.")
                              time.sleep(2.5)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase A Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase A Overcurrent"

                         elif input_A_Phase_Current >= 2 * input_phase_a_over_current_set_value:
                              print("Input Phase A Overcurrent (2x) detected! Waiting 5 sec before trip.")
                              time.sleep(5)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase A Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase A Overcurrent"

                         elif input_A_Phase_Current >= input_phase_a_over_current_set_value:
                              print("Input Phase A Overcurrent detected! Waiting 10 sec before trip.")
                              time.sleep(10)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase A Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase A Overcurrent"

                     except ValueError as e:
                         print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                      # Phase B Overcurrent
              if input_phase_b_over_current_status == 1:
                 if input_B_Phase_Current is not None and input_phase_b_over_current_set_value is not None:
                     try:
                         input_B_Phase_Current = float(input_B_Phase_Current)
                         input_phase_b_over_current_set_value = float(input_phase_b_over_current_set_value)

                         if input_B_Phase_Current >= 20 * input_phase_b_over_current_set_value:
                             print("Input Phase B Overcurrent (20x) detected! Waiting 0.5 sec before trip.")
                             time.sleep(0.5)
                             GPIO.output(input_trip_pin, GPIO.LOW)
                             GPIO.output(led_input, GPIO.HIGH)
                             print("trip on - Input Phase B Overcurrent")
                             input_status = "Unhealthy"
                             breaker_status = "Trip"
                             fault_type = "Input Phase B Overcurrent"

                         elif input_B_Phase_Current >= 10 * input_phase_b_over_current_set_value:
                              print("Input Phase B Overcurrent (10x) detected! Waiting 1 sec before trip.")
                              time.sleep(1)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase B Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase B Overcurrent"

                         elif input_B_Phase_Current >= 5 * input_phase_b_over_current_set_value:
                              print("Input Phase B Overcurrent (5x) detected! Waiting 2.5 sec before trip.")
                              time.sleep(2.5)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase B Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase B Overcurrent"

                         elif input_B_Phase_Current >= 2 * input_phase_b_over_current_set_value:
                              print("Input Phase B Overcurrent (2x) detected! Waiting 5 sec before trip.")
                              time.sleep(5)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase B Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase B Overcurrent"

                         elif input_B_Phase_Current >= input_phase_b_over_current_set_value:
                              print("Input Phase B Overcurrent detected! Waiting 10 sec before trip.")
                              time.sleep(10)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase B Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase B Overcurrent"

                     except ValueError as e:
                         print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                 # Phase C Overcurrent
              if input_phase_c_over_current_status == 1:
                 if input_C_Phase_Current is not None and input_phase_c_over_current_set_value is not None:
                     try:
                         input_C_Phase_Current = float(input_C_Phase_Current)
                         input_phase_c_over_current_set_value = float(input_phase_c_over_current_set_value)

                         if input_C_Phase_Current >= 20 * input_phase_c_over_current_set_value:
                             print("Input Phase C Overcurrent (20x) detected! Waiting 0.5 sec before trip.")
                             time.sleep(0.5)
                             GPIO.output(input_trip_pin, GPIO.LOW)
                             GPIO.output(led_input, GPIO.HIGH)
                             print("trip on - Input Phase C Overcurrent")
                             input_status = "Unhealthy"
                             breaker_status = "Trip"
                             fault_type = "Input Phase C Overcurrent"

                         elif input_C_Phase_Current >= 10 * input_phase_c_over_current_set_value:
                              print("Input Phase C Overcurrent (10x) detected! Waiting 1 sec before trip.")
                              time.sleep(1)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase C Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase C Overcurrent"

                         elif input_C_Phase_Current >= 5 * input_phase_c_over_current_set_value:
                              print("Input Phase C Overcurrent (5x) detected! Waiting 2.5 sec before trip.")
                              time.sleep(2.5)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase C Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase C Overcurrent"

                         elif input_C_Phase_Current >= 2 * input_phase_c_over_current_set_value:
                              print("Input Phase C Overcurrent (2x) detected! Waiting 5 sec before trip.")
                              time.sleep(5)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase C Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase C Overcurrent"

                         elif input_C_Phase_Current >= input_phase_c_over_current_set_value:
                              print("Input Phase C Overcurrent detected! Waiting 10 sec before trip.")
                              time.sleep(10)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input Phase C Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input Phase C Overcurrent"
            

                     except ValueError as e:
                         print(f"Conversion Error: {e} - Check if input values are valid numbers.")


               # Phase A Overcurrent
              if output_phase_a_over_current_status == 1:
                 if output_A_Phase_Current is not None and output_phase_a_over_current_set_value is not None:
                     try:
                         output_A_Phase_Current = float(output_A_Phase_Current)
                         output_phase_a_over_current_set_value = float(output_phase_a_over_current_set_value)

                         if output_A_Phase_Current >= 20 * output_phase_a_over_current_set_value:
                             print("Outpt Phase A Overcurrent (20x) detected! Waiting 0.5 sec before trip.")
                             time.sleep(0.5)
                             GPIO.output(output_trip_pin, GPIO.LOW)
                             GPIO.output(led_output, GPIO.HIGH)
                             print("trip on - Output Phase A Overcurrent")
                             output_status = "Unhealthy"
                             breaker_status = "Trip"
                             fault_type = "Output Phase A Overcurrent"

                         elif output_A_Phase_Current >= 10 * output_phase_a_over_current_set_value:
                              print("Output Phase A Overcurrent (10x) detected! Waiting 1 sec before trip.")
                              time.sleep(1)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase A Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase A Overcurrent"

                         elif output_A_Phase_Current >= 5 * output_phase_a_over_current_set_value:
                              print("Output Phase A Overcurrent (5x) detected! Waiting 2.5 sec before trip.")
                              time.sleep(2.5)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase A Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase A Overcurrent"

                         elif output_A_Phase_Current >= 2 * output_phase_a_over_current_set_value:
                              print("Output Phase A Overcurrent (2x) detected! Waiting 5 sec before trip.")
                              time.sleep(5)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase A Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase A Overcurrent"

                         elif output_A_Phase_Current >= output_phase_a_over_current_set_value:
                              print("Output Phase A Overcurrent detected! Waiting 10 sec before trip.")
                              time.sleep(10)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase A Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase A Overcurrent"
            

                     except ValueError as e:
                         print(f"Conversion Error: {e} - Check if input values are valid numbers.")

                # Phase B Overcurrent
              if output_phase_b_over_current_status == 1:
                 if output_B_Phase_Current is not None and output_phase_b_over_current_set_value is not None:
                     try:
                         output_B_Phase_Current = float(output_B_Phase_Current)
                         output_phase_b_over_current_set_value = float(output_phase_b_over_current_set_value)

                         if output_B_Phase_Current >= 20 * output_phase_b_over_current_set_value:
                             print("Output Phase B Overcurrent (20x) detected! Waiting 0.5 sec before trip.")
                             time.sleep(0.5)
                             GPIO.output(output_trip_pin, GPIO.LOW)
                             GPIO.output(led_output, GPIO.HIGH)
                             print("trip on - Output Phase B Overcurrent")
                             output_status = "Unhealthy"
                             breaker_status = "Trip"
                             fault_type = "Output Phase B Overcurrent"

                         elif output_B_Phase_Current >= 10 * output_phase_b_over_current_set_value:
                              print("Output Phase B Overcurrent (10x) detected! Waiting 1 sec before trip.")
                              time.sleep(1)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase B Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase B Overcurrent"

                         elif output_B_Phase_Current >= 5 * output_phase_b_over_current_set_value:
                              print("Output Phase B Overcurrent (5x) detected! Waiting 2.5 sec before trip.")
                              time.sleep(2.5)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase B Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase B Overcurrent"

                         elif output_B_Phase_Current >= 2 * output_phase_b_over_current_set_value:
                              print("Output Phase B Overcurrent (2x) detected! Waiting 5 sec before trip.")
                              time.sleep(5)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase B Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase B Overcurrent"

                         elif output_B_Phase_Current >= output_phase_b_over_current_set_value:
                              print("Output Phase B Overcurrent detected! Waiting 10 sec before trip.")
                              time.sleep(10)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase B Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase B Overcurrent"
            

                     except ValueError as e:
                         print(f"Conversion Error: {e} - Check if input values are valid numbers.")


               # Phase C Overcurrent
              if output_phase_c_over_current_status == 1:
                 if output_C_Phase_Current is not None and output_phase_c_over_current_set_value is not None:
                     try:
                         output_C_Phase_Current = float(output_C_Phase_Current)
                         output_phase_c_over_current_set_value = float(output_phase_c_over_current_set_value)

                         if output_C_Phase_Current >= 20 * output_phase_c_over_current_set_value:
                             print("Output Phase C Overcurrent (20x) detected! Waiting 0.5 sec before trip.")
                             time.sleep(0.5)
                             GPIO.output(output_trip_pin, GPIO.LOW)
                             GPIO.output(led_output, GPIO.HIGH)
                             print("trip on - Output Phase C Overcurrent")
                             output_status = "Unhealthy"
                             breaker_status = "Trip"
                             fault_type = "Output Phase C Overcurrent"

                         elif output_C_Phase_Current >= 10 * output_phase_c_over_current_set_value:
                              print("Output Phase C Overcurrent (10x) detected! Waiting 1 sec before trip.")
                              time.sleep(1)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase C Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase C Overcurrent"

                         elif output_C_Phase_Current >= 5 * output_phase_c_over_current_set_value:
                              print("Output Phase C Overcurrent (5x) detected! Waiting 2.5 sec before trip.")
                              time.sleep(2.5)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase C Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase C Overcurrent"

                         elif output_C_Phase_Current >= 2 * output_phase_c_over_current_set_value:
                              print("Output Phase C Overcurrent (2x) detected! Waiting 5 sec before trip.")
                              time.sleep(5)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase C Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase C Overcurrent"

                         elif output_C_Phase_Current >= output_phase_c_over_current_set_value:
                              print("Output Phase C Overcurrent detected! Waiting 10 sec before trip.")
                              time.sleep(10)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output Phase C Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output Phase C Overcurrent"

                     except ValueError as e:
                         print(f"Conversion Error: {e} - Check if input values are valid numbers.")
        
              # DC Input Over Current
              if input_dc_over_current_status == 1:
                 if input_DC_Current is not None and input_dc_over_current_set_value is not None:
                     try:
                         input_DC_Current = float(input_DC_Current)
                         input_dc_over_current_set_value = float(input_dc_over_current_set_value)

                         if input_DC_Current >= 20 * input_dc_over_current_set_value:
                             print("Input DC Overcurrent (20x) detected! Waiting 0.5 sec before trip.")
                             time.sleep(0.5)
                             GPIO.output(input_trip_pin, GPIO.LOW)
                             GPIO.output(led_input, GPIO.HIGH)
                             print("trip on - Input DC Overcurrent")
                             input_status = "Unhealthy"
                             breaker_status = "Trip"
                             fault_type = "Input DC Overcurrent"

                         elif input_DC_Current >= 10 * input_dc_over_current_set_value:
                              print("Input DC Overcurrent (10x) detected! Waiting 1 sec before trip.")
                              time.sleep(1)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input DC Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input DC Overcurrent"

                         elif input_DC_Current >= 5 * input_dc_over_current_set_value:
                              print("Input DC Overcurrent (5x) detected! Waiting 2.5 sec before trip.")
                              time.sleep(2.5)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input DC Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input DC Overcurrent"

                         elif input_DC_Current >= 2 * input_dc_over_current_set_value:
                              print("Input DC Overcurrent (2x) detected! Waiting 5 sec before trip.")
                              time.sleep(5)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input DC Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input DC Overcurrent"

                         elif input_DC_Current >= input_dc_over_current_set_value:
                              print("Input DC Overcurrent detected! Waiting 10 sec before trip.")
                              time.sleep(10)
                              GPIO.output(input_trip_pin, GPIO.LOW)
                              GPIO.output(led_input, GPIO.HIGH)
                              print("trip on - Input DC Overcurrent")
                              input_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Input DC Overcurrent"
            

                     except ValueError as e:
                         print(f"Conversion Error: {e} - Check if input values are valid numbers.")

               # DC Output Over Current
              if output_dc_over_current_status == 1:
                 if output_DC_Current is not None and output_dc_over_current_set_value is not None:
                     try:
                         output_DC_Current = float(output_DC_Current)
                         output_dc_over_current_set_value = float(output_dc_over_current_set_value)

                         if output_DC_Current >= 20 * output_dc_over_current_set_value:
                             print("Output DC Overcurrent (20x) detected! Waiting 0.5 sec before trip.")
                             time.sleep(0.5)
                             GPIO.output(output_trip_pin, GPIO.LOW)
                             GPIO.output(led_output, GPIO.HIGH)
                             print("trip on - Output DC Overcurrent")
                             output_status = "Unhealthy"
                             breaker_status = "Trip"
                             fault_type = "Output DC Overcurrent"

                         elif output_DC_Current >= 10 * output_dc_over_current_set_value:
                              print("Output DC Overcurrent (10x) detected! Waiting 1 sec before trip.")
                              time.sleep(1)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output DC Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output DC Overcurrent"

                         elif output_DC_Current >= 5 * output_dc_over_current_set_value:
                              print("Output DC Overcurrent (5x) detected! Waiting 2.5 sec before trip.")
                              time.sleep(2.5)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output DC Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output DC Overcurrent"

                         elif output_DC_Current >= 2 * output_dc_over_current_set_value:
                              print("Output DC Overcurrent (2x) detected! Waiting 5 sec before trip.")
                              time.sleep(5)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output DC Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output DC Overcurrent"

                         elif output_DC_Current >= output_dc_over_current_set_value:
                              print("Output DC Overcurrent detected! Waiting 10 sec before trip.")
                              time.sleep(10)
                              GPIO.output(output_trip_pin, GPIO.LOW)
                              GPIO.output(led_output, GPIO.HIGH)
                              print("trip on - Output DC Overcurrent")
                              output_status = "Unhealthy"
                              breaker_status = "Trip"
                              fault_type = "Output DC Overcurrent"
            
                     except ValueError as e:
                         print(f"Conversion Error: {e} - Check if input values are valid numbers.")
            
                      
                     if fault_type != "None":
                           relay_status = "Unhealthy"
                           breaker_status = "Trip"  # ✅ Ensure breaker status updates
                           print(f"🚨 Fault Detected: {fault_type} - Relay Unhealthy, Breaker Tripped")
   
                            # ✅ Debug print before calling `update_fault_log()`
                           # print(f"✅ Calling update_fault_log() → {relay_status}, {input_status}, {output_status}, {breaker_status}, {fault_type}")
            update_fault_log(relay_status, input_status, output_status, breaker_status, fault_type)
               

            while GPIO.input(input_trip_pin) == GPIO.LOW or GPIO.input(output_trip_pin) == GPIO.LOW:
                 if Reset_button == 1:
                    GPIO.output(input_trip_pin, GPIO.HIGH)
                    GPIO.output(output_trip_pin, GPIO.HIGH)
                    GPIO.output(led_input, GPIO.LOW)
                    GPIO.output(led_output, GPIO.LOW)
                    print("Pins reset ")
                    break
 
                        
            time.sleep(interval)  # Wait before the next update

    except KeyboardInterrupt:
        print("\nMonitoring stopped by user.")

    finally:
        GPIO.cleanup()
        print("GPIO cleaned up.")

# File paths
excel_path = "/home/rahul/Desktop/Project/User Data Input.xlsx"
sheet_name = "Sheet1"
excel_cells = [
    "B2","B3","B4","B5","B6","B7","B8","B9","B10","B11","B12","B13","B14","B15","B16","B17","B18","B19","B20","B21",
    "B22","B23","B24","B25","B26","B27","B28","B29","B30","B31","B32","B33","B34","B35","B36","B37","B38","B39","B40",
    "B41","B42","B43","B44","B45","B46","B47","B48","B49","B50","B51","B52","B53","B54","B55","B56","B57","B58","B59",
    "B60","B61","B62","B63","B64","B65","B66","B67"  
]
 
input_csv = "/home/rahul/Desktop/Project/Input Real Time Data/Real-time data for relay.csv"
output_csv = "/home/rahul/Desktop/Project/Output Real Time Data/Real-time data for relay.csv"
fault_log_file = "/home/rahul/Desktop/Project/Relay_indication.xlsx"
fault_csv_file = "/home/rahul/Desktop/Project/fault_log.csv"

create_fault_log()

# Start monitoring
monitor_files(excel_path, sheet_name, excel_cells, input_csv, output_csv)