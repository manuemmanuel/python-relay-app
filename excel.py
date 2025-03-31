import os
import time
import pandas as pd
from supabase import create_client, Client
from dotenv import load_dotenv
import logging
from openpyxl import load_workbook
import threading

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Initialize Supabase client
supabase_url = os.getenv("SUPABASE_URL")
supabase_key = os.getenv("SUPABASE_KEY")
supabase: Client = create_client(supabase_url, supabase_key)

def update_parameters(xlsx_path):
    try:
        # Get data from Supabase parameters_table
        result = supabase.table('parameters_table').select('*').execute()
        
        if not result.data:
            logger.warning("No data found in parameters_table")
            return
        
        # Create dictionary of parameters and values from Supabase
        supabase_params = {item['parameter'].strip(): item['value'] for item in result.data}
        
        # Load Excel workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        # Track updated parameters with their old and new values
        updates = []
        
        # Skip the header row (row 1) and start from row 2
        for row in range(2, ws.max_row + 1):
            # Get parameter name from column A
            param = ws[f'A{row}'].value
            if param:
                param = param.strip()  # Remove any whitespace
                
                if param in supabase_params:
                    # Get current Excel value and Supabase value
                    excel_value = ws[f'B{row}'].value
                    supabase_value = supabase_params[param]
                    
                    # Convert values to integers for comparison
                    try:
                        excel_value = int(float(excel_value)) if excel_value is not None else 0
                        supabase_value = int(float(supabase_value)) if supabase_value is not None else 0
                        
                        # Update if values are different
                        if excel_value != supabase_value:
                            ws[f'B{row}'].value = supabase_value
                            updates.append({
                                'parameter': param,
                                'old_value': excel_value,
                                'new_value': supabase_value,
                                'row': row
                            })
                    except (ValueError, TypeError) as e:
                        logger.error(f"Error converting values for parameter {param}: {str(e)}")
                        continue
        
        # Save if there were any updates
        if updates:
            wb.save(xlsx_path)
            logger.info("\nUpdated parameters:")
            for update in updates:
                logger.info(f"  Row {update['row']}: {update['parameter']}")
                logger.info(f"    Old value: {update['old_value']}")
                logger.info(f"    New value: {update['new_value']}")
        else:
            logger.info("No updates needed - all parameters are current")
            
    except Exception as e:
        logger.error(f"Error updating parameters: {str(e)}")
        logger.error("Full error details:", exc_info=True)

def update_relay_status(xlsx_path):
    try:
        # Load Excel workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        
        # Get the latest values from Excel
        data = {
            'timestamp': pd.Timestamp.now().isoformat(),
            'relay_status': ws['B1'].value if ws['B1'].value else 'Unknown',
            'input_status': ws['B2'].value if ws['B2'].value else 'Unknown',
            'output_status': ws['B3'].value if ws['B3'].value else 'Unknown',
            'circuit_breaker_status': ws['B4'].value if ws['B4'].value else 'Unknown',
            'fault_type': ws['B5'].value if ws['B5'].value else 'None'
        }

        try:
            # Get existing row if any
            result = supabase.table('device_status').select('id').execute()
            
            if result.data:
                # Update existing row
                row_id = result.data[0]['id']
                supabase.table('device_status').update(data).eq('id', row_id).execute()
                logger.info(f"Updated device status with timestamp: {data['timestamp']}")
            else:
                # Insert new row
                result = supabase.table('device_status').insert(data).execute()
                logger.info(f"Inserted new device status with timestamp: {data['timestamp']}")
                
        except Exception as e:
            logger.error(f"Error updating device_status table: {str(e)}")
            
    except Exception as e:
        logger.error(f"Error processing Relay_indication.xlsx: {str(e)}")

def start_monitoring():
    params_path = "User Data Input.xlsx"
    relay_path = "Relay_indication.xlsx"
    
    logger.info(f"Starting monitoring service for:\n{params_path}\n{relay_path}")
    
    try:
        while True:
            update_parameters(params_path)
            update_relay_status(relay_path)
            time.sleep(1)  # Check every second
    except KeyboardInterrupt:
        logger.info("Stopping monitoring service")

def run_excel_updater():
    excel_thread = threading.Thread(target=start_monitoring, daemon=True)
    excel_thread.start()
    return excel_thread

if __name__ == "__main__":
    start_monitoring()